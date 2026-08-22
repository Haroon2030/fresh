"""استيراد أصناف من ملفات Excel."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from .catalog_units import (
    PACKAGE_NAMES,
    aggregate_catalog_rows,
    merge_unit_strings,
    sanitize_catalog_row,
)

HEADER_ALIASES = {
    'name': {
        'الاسم', 'اسم', 'name', 'item_name', 'الصنف', 'اسمالصنف',
        'اسم الصنف', 'اسم الصنف', 'اسمالصنف',
    },
    'item_number': {
        'رقم الصنف', 'رقم الصنف', 'رقم', 'item_number', 'رقمالصنف',
        'كود', 'code', 'sku',
    },
    'unit': {'الوحدة', 'وحدة', 'unit'},
    'package': {'العبوة', 'عبوة', 'package', 'pack'},
}


def _norm(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _header_key(cell) -> str:
    return _norm(cell).lower().replace(' ', '').replace('_', '')


def _map_headers(row) -> dict[str, int]:
    mapping = {}
    for idx, cell in enumerate(row):
        key = _header_key(cell)
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            normalized_aliases = {a.lower().replace(' ', '').replace('_', '') for a in aliases}
            if key in normalized_aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _detect_layout(header_row) -> dict[str, int] | None:
    """اكتشاف تخطيط ملف طلب الخضار: اسم الصنف | العبوة | رقم الصنف."""
    if not header_row:
        return None
    keys = [_header_key(c) for c in header_row]
    has_name = any(k in {'اسمالصنف', 'الاسم', 'name', 'itemname'} or 'اسمالصن' in k for k in keys)
    has_package = any(k in {'العبوة', 'package', 'pack', 'عبوة'} for k in keys)
    has_number = any('رقمالصن' in k or k in {'رقم', 'itemnumber', 'sku', 'code'} for k in keys)

    if has_name and has_package and has_number and len([k for k in keys if k]) <= 4:
        name_idx = next(i for i, k in enumerate(keys) if k and ('اسمالصن' in k or k in {'الاسم', 'name', 'itemname'}))
        package_idx = next(i for i, k in enumerate(keys) if k in {'العبوة', 'package', 'pack', 'عبوة'})
        number_idx = next(
            i for i, k in enumerate(keys)
            if k and ('رقمالصن' in k or k in {'رقم', 'itemnumber', 'sku', 'code'})
        )
        return {'name': name_idx, 'package': package_idx, 'item_number': number_idx}

    mapped = _map_headers(header_row)
    if 'name' in mapped and 'item_number' in mapped:
        return mapped
    return None


def _looks_like_header(row) -> bool:
    keys = {_header_key(c) for c in (row or ()) if _norm(c)}
    markers = {
        'الاسم', 'name', 'itemname', 'اسمالصنف', 'اسمالصنف',
        'رقمالصنف', 'رقمالصنف', 'itemnumber', 'العبوة', 'package',
    }
    return bool(keys & markers) or any('اسمالصن' in k for k in keys)


def parse_items_workbook(file_obj) -> tuple[list[dict], list[str]]:
    """
    يقرأ ملف Excel ويعيد (صفوف صالحة, أخطاء).
    يدعم: اسم الصنف | العبوة | رقم الصنف (ملف طلب الخضار)
    """
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], ['الملف فارغ.']

    header_row = rows[0]
    header_map = _detect_layout(header_row)
    if header_map is None:
        header_map = {'name': 0, 'item_number': 1, 'unit': 2, 'package': 3}

    data_rows = rows[1:] if _looks_like_header(header_row) else rows

    items = []
    errors = []
    seen_keys = set()
    last_name = ''
    last_number = ''

    for i, row in enumerate(data_rows, start=2 if _looks_like_header(header_row) else 1):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        def cell(field, default=''):
            idx = header_map.get(field)
            if idx is None or idx >= len(row):
                return default
            return _norm(row[idx])

        raw_name = cell('name')
        name = raw_name or last_name
        item_number = cell('item_number') or last_number
        unit = cell('unit')
        package = cell('package')

        name, item_number, package, extra_units = sanitize_catalog_row(
            name,
            package,
            item_number,
            last_name=last_name,
        )
        package = merge_unit_strings(package, *extra_units, unit)

        if not name and not package and not item_number:
            continue
        if name in PACKAGE_NAMES:
            errors.append(f'الصف {i}: «{name}» عبوة وليس اسم صنف — ضع اسم الصنف في الصف الأول.')
            continue
        if not name:
            errors.append(f'الصف {i}: الاسم مطلوب (أو اترك صفاً تحت اسم الصنف لنفس المنتج).')
            continue
        if not package:
            errors.append(f'الصف {i}: العبوة مطلوبة.')
            continue
        if not item_number:
            errors.append(f'الصف {i}: رقم الصنف مطلوب (أو اتركه فارغاً تحت صف له رقم).')
            continue

        dedupe_key = (name.casefold(), package.casefold(), item_number.casefold())
        if dedupe_key in seen_keys:
            errors.append(f'الصف {i}: صف مكرر ({name} / {package}).')
            continue
        seen_keys.add(dedupe_key)

        last_name = name
        last_number = item_number

        items.append({
            'name': name,
            'item_number': item_number,
            'unit': unit,
            'package': package,
        })

    return aggregate_catalog_rows(items), errors


def build_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'الأصناف'
    ws.append(['اسم الصنف', 'العبوة', 'رقم الصنف'])
    ws.append(['خيار', 'جرم', '06142'])
    ws.append(['', 'كيس', ''])
    ws.append(['فلفل شقراء', 'جرم', '06146'])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
